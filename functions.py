import os
import pandas as pd
import re
from openpyxl import load_workbook
from pathlib import Path
from functools import lru_cache
import logging
from logger_utils import log_decorator
from config_manager import config
from typing import Dict

# Использование pathlib для работы с путями
# input_folder = Path(config.INPUT_FOLDER)
# processed_folder = Path(config.PROCESSED_FOLDER)
# log_folder = Path(config.LOG_FOLDER)


@lru_cache(maxsize=10)
def _load_workbook_cached(filepath: str):
    """Кэшированная версия load_workbook для предотвращения многократных открытий одного и того же файла."""
    try:
        return load_workbook(filepath, data_only=True)
    except Exception as e:
        logging.error(f"Ошибка при загрузке книги Excel {filepath}: {e}")
        raise


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Нормализует имена столбцов DataFrame: убирает лишние пробелы и переносы строк,
    сводит несколько пробелов к одному.

    :param df: DataFrame
    :return: DataFrame с нормализованными именами столбцов
    """
    df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip()
    return df


@log_decorator()
def load_named_table(filepath: str, table_name: str) -> pd.DataFrame:
    """
    Загружает таблицу из Excel файла по имени.

    Args:
        filepath (str): Путь к Excel файлу.
        table_name (str): Имя таблицы для загрузки.

    Returns:
        pd.DataFrame: Таблица, загруженная в виде DataFrame.
    """
    try:
        wb = _load_workbook_cached(filepath)
        for sheet in wb.worksheets:
            for tbl in sheet._tables.values():
                if tbl.name == table_name:
                    data = sheet[tbl.ref]
                    rows = [[cell.value for cell in row] for row in data]
                    # создаем DataFrame
                    df = pd.DataFrame(rows[1:], columns=rows[0])
                    # Применяем нормализацию столбцов
                    return normalize_columns(df)
        raise ValueError(f"Таблица '{table_name}' не найдена")
    except Exception as e:
        raise RuntimeError(f"Ошибка загрузки: {e}")


@log_decorator()
def load_all_tables_from_file(filepath: str, verbose: bool = False) -> list[pd.DataFrame]:
    tables = []
    try:
        wb = _load_workbook_cached(filepath)
        all_table_names = []

        for sheet in wb.worksheets:
            for tbl in sheet._tables.values():
                all_table_names.append(tbl.name)
                data = sheet[tbl.ref]
                rows = [[cell.value for cell in row] for row in data]
                headers = rows[0]
                values = rows[1:]
                df = pd.DataFrame(values, columns=headers)
                tables.append(df)

        msg = f"В файле '{filepath}' найдены таблицы: \n{all_table_names}"
        if verbose:
            print(msg)
        logging.info(msg)
        return tables
    except Exception as e:
        raise RuntimeError(
            f"Ошибка при чтении всех таблиц из файла {filepath}: {e}")


@log_decorator()
def combine_dataframes(dfs, columns_to_remove, rename_map):
    combined = pd.concat(dfs, ignore_index=True)
    combined = combined.copy()
    combined.drop(columns=[
                  col for col in columns_to_remove if col in combined.columns], errors='ignore', inplace=True)
    combined.rename(columns=rename_map, inplace=True)
    return combined


@log_decorator()
def save_dataframe_to_excel(df: pd.DataFrame, path: str, verbose=False) -> None:
    df.to_excel(path, index=False)
    msg = f"Результат сохранён: {path}"
    if verbose:
        print(msg)
    logging.info(msg)


@log_decorator()
def smart_merge(df: pd.DataFrame, rename_map: dict[str, str]) -> pd.DataFrame:
    """
    Удаляет дубликаты по комбинации столбцов 'ФИО' и 'УЗ', сохраняет первое вхождение как оригинал,
    заполняет пропуски у оригинала значениями из дубликатов, удаляет дубли.

    Args:
        df: Входной DataFrame
        rename_map: Словарь для переименования столбцов

    Returns:
        DataFrame без дубликатов с заполненными пропусками
    """
    df = df.copy()
    df.rename(columns=rename_map, inplace=True)

    # Группируем по 'ФИО' и 'УЗ'
    grouped = df.groupby(['ФИО', 'УЗ'], as_index=False)

    # Для каждой группы обрабатываем дубликаты
    final_rows = []

    for _, group in grouped:
        # Оставляем первую строку как оригинал
        original = group.iloc[0].copy()

        # Проходим по остальным строкам в группе (дубликаты)
        for _, duplicate in group.iloc[1:].iterrows():
            # Если у оригинала отсутствуют значения, подставляем их из дубликатов
            for col in df.columns:
                if pd.isna(original[col]) and not pd.isna(duplicate[col]):
                    original[col] = duplicate[col]

        # Добавляем обработанную строку в финальный список
        final_rows.append(original)

    # Создаем новый DataFrame из обработанных строк
    final_df = pd.DataFrame(final_rows)

    return final_df


@log_decorator()
def apply_replacements(df, replace_dict):
    """
    Заменяет значения в DataFrame по указанному словарю замен.
    Убирает лишние пробелы и переносы строк в именах столбцов перед применением замен.

    :param df: DataFrame, с данными
    :param replace_dict: словарь замен в формате {столбец: {старое значение: новое значение}}
    :return: DataFrame с применёнными заменами
    """

    # Убираем лишние пробелы и переносы строк из имен столбцов
    df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip()

    # Проходим по всем столбцам и заменяем их значения
    for column, replacements in replace_dict.items():
        if column in df.columns:
            for old_value, new_value in replacements.items():
                # Заменяем значения в столбце
                df[column] = df[column].replace(old_value, new_value)

    return df


@log_decorator()
def combine_columns_by_replace_key(df: pd.DataFrame, replace_key: str, config, drop: bool = False) -> pd.DataFrame:
    """
    Объединяет значения из указанных столбцов (по ключу replace_key из config) для каждой строки.
    Если в config есть блок "+", его значение будет использовано как замена.
    Итоговое значение объединяется через '!' и записывается обратно в столбец (если он один) или можно расширить логику.

    Пример:
    Для replace_key = "REPLACE_ENERGYMAIN":
    {
        "Устранение дефектов": { "+": "uuid1" },
        "Ответственный за устранение дефектов": { "+": "uuid2" }
    }

    Значения из этих двух столбцов будут объединены через "!".
    """
    replace_config = getattr(config, replace_key, {})
    columns_to_combine = list(replace_config.keys())

    if not columns_to_combine:
        return df

    # Объединяем значения по строкам
    def combine_row_values(row):
        values = []
        for col in columns_to_combine:
            val = row.get(col)
            if pd.notna(val) and str(val).strip():
                values.append(str(val).strip())
        return "!".join(values)

    # Новый столбец, по replace_key
    new_column_name = f"{replace_key}_combined"

    df[new_column_name] = df.apply(combine_row_values, axis=1)

    # Удаление столбцов, которые были объединены
    if drop:
        for col in replace_config.keys():
            if col in df.columns:
                logging.debug(
                    f"Удаляем столбец '{col}', так как он был объединен.")
                df.drop(columns=[col], inplace=True)

    return df
